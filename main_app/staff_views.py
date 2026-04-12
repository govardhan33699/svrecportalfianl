import json
import pandas as pd
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
from PIL import Image as PilImage

from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt

# Helper functions for Excel Parity
def js_round(v):
    """Matches JavaScript Math.round() for positive numbers."""
    try:
        return int(float(v) + 0.5)
    except (ValueError, TypeError):
        return 0

def fmt_qty(v):
    """Formats 6.00 as 6, and 24.85 as 24.85."""
    try:
        v = float(v)
        if v.is_integer():
            return int(v)
        return round(v, 2)
    except (ValueError, TypeError):
        return 0
from django.shortcuts import HttpResponseRedirect, get_object_or_404, redirect, render

from .models import (
    Staff, Student, Subject, Attendance, AttendanceReport, LeaveReportStaff,
    StudentResult, Session, Course, Period, Timetable, Announcement,
    NotificationStaff, Book, IssuedBook, Assignment, AssignmentSubmission, FeedbackStaff, StudyMaterial,
    SEMESTER_CHOICES
)
from .forms import (
    StaffEditForm, LeaveReportStaffForm, FeedbackStaffForm, AssignmentForm,
    StaffQualificationFormSet
)
from datetime import date

def staff_home(request):
    staff = get_object_or_404(Staff, admin=request.user)
    total_students = Student.objects.filter(course=staff.course).count()
    total_leave = LeaveReportStaff.objects.filter(staff=staff).count()
    from django.db.models import Count
    subjects = Subject.objects.filter(staff=staff).annotate(attendance_count=Count('attendance'))
    total_subject = subjects.count()
    attendance_list = Attendance.objects.filter(subject__in=subjects)
    total_attendance = attendance_list.count()
    attendance_list = []
    subject_list = []
    for subject in subjects:
        subject_list.append(subject.name)
        attendance_list.append(subject.attendance_count)

    context = {
        'page_title': 'Staff Panel - ' + str(staff.admin.first_name) + ' ' + str(staff.admin.last_name[0]) + '' + ' (' + str(staff.course) + ')',
        'total_students': total_students,
        'total_attendance': total_attendance,
        'total_leave': total_leave,
        'total_subject': total_subject,
        'subject_list': subject_list,
        'attendance_list': attendance_list
    }

    # Timetable for today (staff-specific)
    from datetime import date as dt_date
    today_name = dt_date.today().strftime('%A')
    timetable_today = Timetable.objects.filter(
        staff=staff, day=today_name
    ).select_related('period', 'subject', 'course').order_by('period__number')
    context['today_name'] = today_name
    context['timetable_today'] = timetable_today

    # Recent Activities for Staff
    recent_attendance = Attendance.objects.filter(subject__staff=staff).select_related('subject').order_by("-created_at")[:3]
    recent_marks = StudentResult.objects.filter(subject__staff=staff).select_related('student__admin', 'subject').order_by("-updated_at")[:3]
    recent_leaves = LeaveReportStaff.objects.filter(staff=staff).order_by("-created_at")[:3]


    activities = []
    for att in recent_attendance:
        activities.append({
            'title': f"Attendance Taken: {att.subject.name}",
            'time': att.created_at,
            'icon': 'fa-check-circle',
            'color': 'success'
        })
    for mark in recent_marks:
        activities.append({
            'title': f"Marks Added: {mark.student.admin.get_full_name()} ({mark.subject.name})",
            'time': mark.updated_at,
            'icon': 'fa-edit',
            'color': 'primary'
        })
    for leave in recent_leaves:
        status = "Pending" if leave.status == 0 else ("Approved" if leave.status == 1 else "Rejected")
        activities.append({
            'title': f"Leave Request {status}: {leave.date}",
            'time': leave.created_at,
            'icon': 'fa-calendar-minus',
            'color': 'warning' if leave.status == 0 else ('success' if leave.status == 1 else 'danger')
        })

    activities.sort(key=lambda x: x['time'], reverse=True)
    context['recent_activities'] = activities[:10]
    return render(request, "staff_template/erpnext_staff_home.html", context)


def staff_take_attendance(request):
    staff = get_object_or_404(Staff, admin=request.user)
    subjects = Subject.objects.filter(staff_id=staff)
    sessions = Session.objects.all()
    courses = Course.objects.all()
    sections = Student.SECTION
    semesters = SEMESTER_CHOICES
    
    context = {
        'subjects': subjects,
        'sessions': sessions,
        'courses': courses,
        'sections': sections,
        'semesters': semesters,
        'page_title': 'Take Attendance'
    }

    return render(request, 'staff_template/staff_take_attendance.html', context)

@csrf_exempt
def get_faculty_attendance_grid(request):
    # ✅ Add role validation
    if request.user.user_type != '2':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    staff = get_object_or_404(Staff, admin=request.user)
    date_str = request.POST.get('date')
    
    try:
        from datetime import datetime
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        day_name = date_obj.strftime('%A')
        
        periods = Period.objects.all().order_by('number')
        timetable = Timetable.objects.filter(staff=staff, day=day_name).select_related('subject', 'course', 'period')
        
        # Attendance already taken today for this staff
        attendance_taken = Attendance.objects.filter(date=date_obj, subject__staff=staff).select_related('subject', 'period')
        taken_map = {(a.subject_id, a.period_id): True for a in attendance_taken}
        
        grid_data = []
        for p in periods:
            # Find timetable entry for this period
            entry = timetable.filter(period=p).first()
            status = "no_class"
            subject_name = ""
            subject_id = ""
            course_name = ""
            section = ""
            
            if entry:
                subject_name = entry.subject.name
                subject_id = entry.subject.id
                course_name = entry.course.name
                section = entry.section
                if taken_map.get((entry.subject_id, p.id)):
                    status = "completed"
                else:
                    status = "pending"
            
            grid_data.append({
                'period_id': p.id,
                'period_number': p.number,
                'start_time': p.start_time.strftime('%H:%M'),
                'end_time': p.end_time.strftime('%H:%M'),
                'status': status,
                'subject_name': subject_name,
                'subject_id': subject_id,
                'course_name': course_name,
                'section': section
            })
            
        return JsonResponse(grid_data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@csrf_exempt
def get_students(request):
    try:
        subject_id = request.POST.get('subject')
        session_id = request.POST.get('session')
        course_idSelection = request.POST.get('branch')
        section = request.POST.get('section')
        semester = request.POST.get('semester')
        attendance_date = request.POST.get('date')  # Optional
        period_id = request.POST.get('period')     # Optional
        exam_name = request.POST.get('exam_name')   # Optional, for marks entry
        
        if not all([subject_id, session_id, section]):
            return JsonResponse({"error": "Missing mandatory fields (Subject, Session, Section)"}, status=400)

        if not course_idSelection:
            subject = get_object_or_404(Subject, id=subject_id)
            course_idSelection = subject.course.id

        students_query = Student.objects.filter(
            course_id=course_idSelection, 
            session_id=session_id,
            section=section
        )

        if semester:
            students_query = students_query.filter(semester=semester)
        
        students = students_query.all()

        # Check for existing attendance if date and period are provided
        attendance_map = {}
        if attendance_date and period_id:
            attendance = Attendance.objects.filter(date=attendance_date, subject_id=subject_id, period_id=period_id).first()
            if attendance:
                reports = AttendanceReport.objects.filter(attendance=attendance)
                attendance_map = {r.student_id: r.status for r in reports}

        # Check for existing marks if exam_name and subject_id are provided
        marks_map = {}
        mid1_map = {}
        mid2_map = {}
        
        if exam_name and subject_id:
            if exam_name == "Final Internal":
                # Fetch both Mid 1 and Mid 2
                mid1_results = StudentResult.objects.filter(subject_id=subject_id, exam_name='Mid 1')
                mid2_results = StudentResult.objects.filter(subject_id=subject_id, exam_name='Mid 2')
                
                if semester:
                    mid1_results = mid1_results.filter(semester=semester)
                    mid2_results = mid2_results.filter(semester=semester)
                
                # Order by ID to ensure .last() behavior is consistent if overlapping records exist
                mid1_results = mid1_results.order_by('id')
                mid2_results = mid2_results.order_by('id')

                for r in mid1_results:
                    mid1_map[r.student_id] = {'obj': r.objective, 'dis': r.descriptive, 'asgn': r.assignment, 'total': r.total}
                for r in mid2_results:
                    mid2_map[r.student_id] = {'obj': r.objective, 'dis': r.descriptive, 'asgn': r.assignment, 'total': r.total}
                
                # Current Final Internal records if any
                main_results = StudentResult.objects.filter(subject_id=subject_id, exam_name=exam_name)
                for r in main_results:
                    marks_map[r.student_id] = {'total': r.total, 'exists': True}
            else:
                results = StudentResult.objects.filter(subject_id=subject_id, exam_name=exam_name)
                for r in results:
                    marks_map[r.student_id] = {
                        'objective': r.objective,
                        'descriptive': r.descriptive,
                        'assignment': r.assignment,
                        'total': r.total,
                        'exists': True
                    }
        
        student_data = []
        for student in students:
            # Default to True (Present) unless a record says otherwise
            status = attendance_map.get(student.id, True)
            
            if exam_name == "Final Internal":
                m1 = mid1_map.get(student.id, {'obj': 0, 'dis': 0, 'asgn': 0, 'total': 0})
                m2 = mid2_map.get(student.id, {'obj': 0, 'dis': 0, 'asgn': 0, 'total': 0})
                final_marks = marks_map.get(student.id, {'total': 0, 'exists': False})
                
                data = {
                    "id": student.id,
                    "name": f"{student.admin.last_name} {student.admin.first_name}".strip(),
                    "roll_number": student.roll_number or "",
                    "attendance_status": status,
                    "mid1": m1,
                    "mid2": m2,
                    "marks": final_marks, # Final Internal only tracks total
                    "profile_pic": student.admin.profile_pic.url if student.admin.profile_pic else ""
                }
            else:
                marks = marks_map.get(student.id, {'objective': 0, 'descriptive': 0, 'assignment': 0, 'total': 0, 'exists': False})
                data = {
                    "id": student.id,
                    "name": f"{student.admin.last_name} {student.admin.first_name}".strip(),
                    "roll_number": student.roll_number or "",
                    "attendance_status": status,
                    "marks": marks,
                    "profile_pic": student.admin.profile_pic.url if student.admin.profile_pic else ""
                }
            student_data.append(data)
            
        return JsonResponse(student_data, safe=False)
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def save_attendance(request):
    student_data = request.POST.get('student_ids')
    date = request.POST.get('date')
    subject_id = request.POST.get('subject')
    session_id = request.POST.get('session')
    period_id = request.POST.get('period')
    students = json.loads(student_data)
    try:
        session = get_object_or_404(Session, id=session_id)
        subject = get_object_or_404(Subject, id=subject_id)
        period = get_object_or_404(Period, id=period_id) if period_id else None
        
        # Check if attendance already exists for this period/date/subject
        if Attendance.objects.filter(date=date, subject=subject, period=period).exists():
            return HttpResponse("ALREADY_TAKEN")

        # Get the semester from the first student to categorize this attendance record
        first_student = get_object_or_404(Student, id=students[0].get('id'))
        semester = first_student.semester

        attendance = Attendance(session=session, subject=subject, date=date, period=period, semester=semester)
        attendance.save()

        for student_dict in students:
            student = get_object_or_404(Student, id=student_dict.get('id'))
            attendance_report = AttendanceReport(student=student, attendance=attendance, status=student_dict.get('status'))
            attendance_report.save()
    except Exception as e:
        return HttpResponse(str(e))

    return HttpResponse("OK")


def staff_update_attendance(request):
    staff = get_object_or_404(Staff, admin=request.user)
    subjects = Subject.objects.filter(staff_id=staff)
    sessions = Session.objects.all()
    context = {
        'subjects': subjects,
        'sessions': sessions,
        'page_title': 'Update Attendance'
    }

    return render(request, 'staff_template/staff_update_attendance.html', context)


def get_student_attendance(request):
    attendance_date_id = request.POST.get('attendance_date_id')
    try:
        date = get_object_or_404(Attendance, id=attendance_date_id)
        attendance_data = AttendanceReport.objects.filter(attendance=date)
        student_data = []
        for attendance in attendance_data:
            data = {"id": attendance.student.admin.id,
                    "name": attendance.student.admin.last_name + " " + attendance.student.admin.first_name,
                    "roll_number": attendance.student.roll_number,
                    "status": attendance.status}
            student_data.append(data)
        return JsonResponse(json.dumps(student_data), content_type='application/json', safe=False)
    except Exception as e:
        return e


def update_attendance(request):
    student_data = request.POST.get('student_ids')
    date = request.POST.get('date')
    students = json.loads(student_data)
    try:
        attendance = get_object_or_404(Attendance, id=date)

        for student_dict in students:
            student = get_object_or_404(
                Student, admin_id=student_dict.get('id'))
            attendance_report = get_object_or_404(AttendanceReport, student=student, attendance=attendance)
            attendance_report.status = student_dict.get('status')
            attendance_report.save()
    except Exception as e:
        return None

    return HttpResponse("OK")


def staff_apply_leave(request):
    form = LeaveReportStaffForm(request.POST or None)
    staff = get_object_or_404(Staff, admin_id=request.user.id)
    context = {
        'form': form,
        'leave_history': LeaveReportStaff.objects.filter(staff=staff),
        'page_title': 'Apply for Leave'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.staff = staff
                obj.save()
                messages.success(
                    request, "Application for leave has been submitted for review")
                return redirect(reverse('staff_apply_leave'))
            except Exception:
                messages.error(request, "Could not apply!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "staff_template/staff_apply_leave.html", context)


def staff_feedback(request):
    form = FeedbackStaffForm(request.POST or None)
    staff = get_object_or_404(Staff, admin_id=request.user.id)
    context = {
        'form': form,
        'feedbacks': FeedbackStaff.objects.filter(staff=staff),
        'page_title': 'Add Feedback'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.staff = staff
                obj.save()
                messages.success(request, "Feedback submitted for review")
                return redirect(reverse('staff_feedback'))
            except Exception:
                messages.error(request, "Could not Submit!")
        else:
            messages.error(request, "Form has errors!")
    return render(request, "staff_template/staff_feedback.html", context)


def staff_view_profile(request):
    staff = get_object_or_404(Staff, admin=request.user)
    form = StaffEditForm(request.POST or None, request.FILES or None,instance=staff)
    formset = StaffQualificationFormSet(request.POST or None, instance=staff)
    context = {'form': form, 'formset': formset, 'page_title': 'View/Update Profile'}
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                address = form.cleaned_data.get('address')
                gender = form.cleaned_data.get('gender')
                passport = request.FILES.get('profile_pic') or None
                admin = staff.admin
                if password != None:
                    admin.set_password(password)
                if passport != None:
                    admin.profile_pic = passport
                admin.first_name = first_name
                admin.last_name = last_name
                admin.address = address
                admin.gender = gender
                admin.save()
                
                staff.course = form.cleaned_data.get('course')
                staff.designation = form.cleaned_data.get('designation')
                staff.date_of_birth = form.cleaned_data.get('date_of_birth')
                staff.date_of_joining = form.cleaned_data.get('date_of_joining')
                staff.experience = form.cleaned_data.get('experience')
                staff.mobile_number = form.cleaned_data.get('mobile_number')
                staff.father_name = form.cleaned_data.get('father_name')
                staff.mother_name = form.cleaned_data.get('mother_name')
                staff.aadhaar_number = form.cleaned_data.get('aadhaar_number')
                staff.pan_number = form.cleaned_data.get('pan_number')
                staff.spouse_name = form.cleaned_data.get('spouse_name')
                staff.qualification = form.cleaned_data.get('qualification')
                staff.blood_group = form.cleaned_data.get('blood_group')
                staff.faculty_role = form.cleaned_data.get('faculty_role')
                staff.save()
                
                if formset.is_valid():
                    formset.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('staff_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
                return render(request, "staff_template/staff_view_profile.html", context)
        except Exception as e:
            messages.error(
                request, "Error Occured While Updating Profile " + str(e))
            return render(request, "staff_template/staff_view_profile.html", context)

    return render(request, "staff_template/staff_view_profile.html", context)


def staff_fcmtoken(request):
    token = request.POST.get('token')
    try:
        staff_user = get_object_or_404(CustomUser, id=request.user.id)
        staff_user.fcm_token = token
        staff_user.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def staff_view_notification(request):
    staff = get_object_or_404(Staff, admin=request.user)
    notifications = NotificationStaff.objects.filter(staff=staff)
    context = {
        'notifications': notifications,
        'page_title': "View Notifications"
    }
    return render(request, "staff_template/staff_view_notification.html", context)


def staff_add_result(request):
    # ✅ FIX BUG #4: Correct role check - staff view should only allow staff type '2'
    if request.user.user_type == '2':  # Staff member
        staff = get_object_or_404(Staff, admin=request.user)
        subjects = Subject.objects.filter(staff=staff)
    else:
        # Non-staff users should not access this view
        messages.error(request, "Unauthorized access!")
        return redirect(reverse('login_page'))
    
    courses = Course.objects.all()
    sections = Student.SECTION
    semesters = SEMESTER_CHOICES
    exam_choices = ["Mid 1", "Mid 2", "Final Internal"]
    sessions = Session.objects.all()
    
    context = {
        'page_title': 'Marks Upload',
        'subjects': subjects,
        'sessions': sessions,
        'courses': courses,
        'sections': sections,
        'semesters': semesters,
        'exam_choices': exam_choices
    }
    if request.method == 'POST':
        try:
            student_ids = request.POST.getlist('student_ids[]')
            subject_id = request.POST.get('subject')
            exam_name = request.POST.get('exam_name')
            subject = get_object_or_404(Subject, id=subject_id)

            for student_id in student_ids:
                objective = request.POST.get('objective_' + student_id, 0)
                descriptive = request.POST.get('descriptive_' + student_id, 0)
                assignment = request.POST.get('assignment_' + student_id, 0)
                
                # Convert to float and calculate total
                try:
                    objective = float(objective)
                    descriptive = float(descriptive)
                    assignment = float(assignment)
                except (ValueError, TypeError):
                    objective = descriptive = assignment = 0
                
                # Enforce limits
                objective = min(objective, 10)
                descriptive = min(descriptive, 30)
                assignment = min(assignment, 5)
                
                total = min(objective + (descriptive / 2) + assignment, 30)
                student = get_object_or_404(Student, id=student_id)
                
                try:
                    data = StudentResult.objects.get(
                        student=student, subject=subject, exam_name=exam_name, semester=subject.semester)
                    data.objective = objective
                    data.descriptive = descriptive
                    data.assignment = assignment
                    data.total = total
                    data.save()
                except StudentResult.DoesNotExist:
                    result = StudentResult(student=student, subject=subject, objective=objective, descriptive=descriptive, assignment=assignment, total=total, exam_name=exam_name, semester=subject.semester)
                    result.save()
            
            messages.success(request, "Marks Saved Successfully")
        except Exception as e:
            messages.warning(request, "Error Occured While Processing Form: " + str(e))
    return render(request, "staff_template/staff_add_result.html", context)


def fetch_student_result(request):
    try:
        subject_id = request.POST.get('subject')
        student_id = request.POST.get('student')
        student = get_object_or_404(Student, id=student_id)
        subject = get_object_or_404(Subject, id=subject_id)
        result = StudentResult.objects.filter(student=student, subject=subject).last()
        if not result:
            return HttpResponse('False')
            
        result_data = {
            'objective': result.objective,
            'descriptive': result.descriptive,
            'assignment': result.assignment,
            'total': result.total
        }
        return JsonResponse(result_data, safe=False)
    except Exception as e:
        return HttpResponse('False')

def export_marks_template(request):
    try:
        subject_id = request.GET.get('subject')
        session_id = request.GET.get('session')
        course_id = request.GET.get('branch')
        section = request.GET.get('section')
        semester = request.GET.get('semester')

        if not all([subject_id, session_id, section, course_id, semester]):
            return HttpResponse("Missing filter parameters")

        students = Student.objects.filter(
            course_id=course_id,
            session_id=session_id,
            section=section,
            semester=semester
        ).select_related('admin')

        data = []
        for s in students:
            data.append({
                'Student ID': s.id,
                'Roll Number': s.roll_number,
                'Name': f"{s.admin.last_name} {s.admin.first_name}".strip(),
                'Descriptive (30)': 0,
                'Objective (10)': 0,
                'Assignment (5)': 0
            })

        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Marks')
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=marks_template_{section}_{semester}.xlsx'
        return response
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")

def download_generic_template(request):
    try:
        data = [{
            'Student ID': '',
            'Descriptive (30)': '',
            'Objective (10)': '',
            'Assignment (5)': ''
        }]
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Marks')
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=sample_marks_template.xlsx'
        return response
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")

def import_marks_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            subject_id = request.POST.get('subject')
            exam_name = request.POST.get('exam_name')
            
            if not all([subject_id, exam_name]):
                messages.error(request, "Please select Subject and Exam Name first.")
                return redirect(reverse('staff_add_result'))

            subject = get_object_or_404(Subject, id=subject_id)
            df = pd.read_excel(excel_file)
            
            # Required columns validation
            required_cols = ['Student ID', 'Descriptive (30)', 'Objective (10)', 'Assignment (5)']
            if not all(col in df.columns for col in required_cols):
                messages.error(request, f"Excel file must contain columns: {', '.join(required_cols)}")
                return redirect(reverse('staff_add_result'))

            count = 0
            for index, row in df.iterrows():
                try:
                    student_id = row['Student ID']
                    desc = float(row.get('Descriptive (30)', 0))
                    obj = float(row.get('Objective (10)', 0))
                    assign = float(row.get('Assignment (5)', 0))
                    
                    # Limits
                    obj = min(max(obj, 0), 10)
                    desc = min(max(desc, 0), 30)
                    assign = min(max(assign, 0), 5)
                    
                    total = obj + (desc / 2) + assign
                    student = get_object_or_404(Student, id=student_id)
                    
                    # Update or create
                    StudentResult.objects.update_or_create(
                        student=student,
                        subject=subject,
                        exam_name=exam_name,
                        semester=subject.semester,
                        defaults={
                            'objective': obj,
                            'descriptive': desc,
                            'assignment': assign,
                            'total': total
                        }
                    )
                    count += 1
                except Exception as row_error:
                    print(f"Error processing row {index}: {row_error}")
                    continue

            messages.success(request, f"Successfully uploaded marks for {count} students.")
        except Exception as e:
            messages.error(request, f"Error processing Excel: {str(e)}")
    else:
        messages.error(request, "No file uploaded.")
    
    return redirect(reverse('staff_add_result'))

#materials
def staff_add_material(request):
    staff = get_object_or_404(Staff, admin=request.user)
    subjects = Subject.objects.filter(staff=staff)
    materials = StudyMaterial.objects.filter(subject__in=subjects).order_by('-created_at')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        subject_id = request.POST.get('subject')
        file = request.FILES.get('file')
        description = request.POST.get('description')
        
        if title and subject_id and file:
            try:
                subject = get_object_or_404(Subject, id=subject_id)
                material = StudyMaterial.objects.create(
                    title=title,
                    subject=subject,
                    file=file,
                    description=description
                )
                messages.success(request, "Material uploaded successfully!")
                return redirect(reverse('staff_add_material'))
            except Exception as e:
                messages.error(request, f"Error uploading material: {str(e)}")
        else:
            messages.error(request, "Please fill all required fields.")
            
    context = {
        'page_title': 'Manage Study Materials',
        'subjects': subjects,
        'materials': materials
    }
    return render(request, "staff_template/staff_add_material.html", context)


def staff_create_assignment(request):
    staff = get_object_or_404(Staff, admin=request.user)
    subjects = Subject.objects.filter(staff=staff)
    if request.method == 'POST':
        form = AssignmentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                assignment = form.save(commit=False)
                assignment.staff = staff
                assignment.save()
                messages.success(request, "Assignment Created Successfully")
                return redirect(reverse('staff_manage_assignments'))
            except Exception as e:
                messages.error(request, "Could Not Create Assignment: " + str(e))
        else:
            messages.error(request, "Please fill the form properly")
    else:
        form = AssignmentForm()
    # Filter subjects to only staff's subjects
    form.fields['subject'].queryset = subjects
    context = {
        'form': form,
        'page_title': 'Create Assignment'
    }
    return render(request, "staff_template/staff_create_assignment.html", context)


def staff_manage_assignments(request):
    staff = get_object_or_404(Staff, admin=request.user)
    assignments = Assignment.objects.filter(staff=staff).order_by('-created_at')
    context = {
        'assignments': assignments,
        'page_title': 'Manage Assignments'
    }
    return render(request, "staff_template/staff_manage_assignments.html", context)


def staff_view_submissions(request, assignment_id):
    staff = get_object_or_404(Staff, admin=request.user)
    assignment = get_object_or_404(Assignment, id=assignment_id, staff=staff)
    submissions = AssignmentSubmission.objects.filter(assignment=assignment)
    context = {
        'assignment': assignment,
        'submissions': submissions,
        'page_title': f'Submissions - {assignment.title}'
    }
    return render(request, "staff_template/staff_view_submissions.html", context)


@csrf_exempt
def staff_grade_submission(request):
    if request.method == 'POST':
        submission_id = request.POST.get('submission_id')
        marks = request.POST.get('marks')
        remarks = request.POST.get('remarks')
        status = request.POST.get('status')

        try:
            submission = get_object_or_404(AssignmentSubmission, id=submission_id)
            
            if marks:
                try:
                    marks_val = float(marks)
                    if marks_val < 0 or marks_val > 5:
                        return JsonResponse({"status": "error", "message": "Marks must be between 0 and 5"})
                    submission.marks = marks_val
                except ValueError:
                    return JsonResponse({"status": "error", "message": "Invalid marks format"})
            else:
                submission.marks = None

            submission.remarks = remarks
            submission.status = status
            submission.save()
            return JsonResponse({"status": "success", "message": "Submission Graded Successfully"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
    return JsonResponse({"status": "error", "message": "Invalid Request"})


def staff_delete_assignment(request, assignment_id):
    staff = get_object_or_404(Staff, admin=request.user)
    assignment = get_object_or_404(Assignment, id=assignment_id, staff=staff)
    assignment.delete()
    messages.success(request, "Assignment Deleted Successfully")
    return redirect(reverse('staff_manage_assignments'))


def staff_add_timetable(request):
    staff = get_object_or_404(Staff, admin=request.user)
    form = StaffTimetableForm(request.POST or None, staff_id=request.user.id)
    context = {
        'form': form,
        'page_title': 'Add Timetable'
    }
    if request.method == 'POST':
        if form.is_valid():
            try:
                obj = form.save(commit=False)
                obj.staff = staff
                obj.save()
                messages.success(request, "Timetable Entry Added Successfully")
                return redirect(reverse('staff_manage_timetable'))
            except Exception as e:
                messages.error(request, "Could not add entry: " + str(e))
        else:
            messages.error(request, "Form has errors")
    return render(request, "staff_template/staff_add_timetable.html", context)


def staff_manage_timetable(request):
    staff = get_object_or_404(Staff, admin=request.user)
    timetables = Timetable.objects.filter(staff=staff).select_related('period', 'course', 'subject')
    
    from collections import defaultdict
    DAY_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    periods = list(Period.objects.all().order_by('number'))

    # Build lookup: {day: {period_number: entry}}
    lookup = defaultdict(dict)
    for entry in timetables:
        lookup[entry.day][entry.period.number] = entry

    # Build grid rows: [{day, cells: [entry or None, ...]}]
    grid_rows = []
    for day in DAY_ORDER:
        cells = [lookup[day].get(p.number) for p in periods]
        grid_rows.append({'day': day, 'cells': cells})

    context = {
        'periods': periods,
        'grid_rows': grid_rows,
        'page_title': 'My Timetable'
    }
    return render(request, "staff_template/staff_manage_timetable.html", context)


def staff_view_announcement(request):
    from django.db.models import Q
    today = date.today()
    announcements = Announcement.objects.filter(
        Q(audience__in=['all', 'staff']) & 
        (Q(expires_at__isnull=True) | Q(expires_at__gte=today))
    ).order_by('-created_at')
    
    context = {
        'announcements': announcements,
        'page_title': 'Announcements'
    }
    return render(request, "staff_template/staff_view_announcement.html", context)
from django.conf import settings

def export_final_internal(request):
    try:
        subject_id = request.GET.get('subject')
        session_id = request.GET.get('session')
        section = request.GET.get('section')
        semester = request.GET.get('semester')
        
        if not all([subject_id, session_id, section, semester]):
            return HttpResponse("Missing parameters for Excel export")

        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Session, id=session_id)
        
        students = Student.objects.filter(
            course=subject.course,
            session=session,
            section=section,
            semester=semester
        ).order_by('roll_number')

        wb = Workbook()
        ws = wb.active
        ws.title = "Final Internal Marks"

        # Styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
        sub_header_fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")
        title_font = Font(bold=True, size=15, color="1a3c5e")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )

        # Branding Header (Rows 1-4)
        ws.merge_cells('A1:A4') # Logo Space
        ws.merge_cells('B1:Q1') # College Name
        ws.merge_cells('B2:Q2') # Affiliation
        ws.merge_cells('B3:Q3') # Address
        ws.merge_cells('B4:Q4') # Department info or spacing

        ws['B1'] = "SVR ENGINEERING COLLEGE"
        ws['B1'].font = Font(bold=True, size=20, color="d32f2f") # Red for emphasis
        ws['B1'].alignment = center_align

        ws['B2'] = "(Approved by AICTE, New Delhi, Affiliated to JNTUA, Ananthapuramu)"
        ws['B2'].font = Font(bold=True, size=10)
        ws['B2'].alignment = center_align

        ws['B3'] = "Ayyalur Metta, Nandyal, Andhra Pradesh - 518501"
        ws['B3'].font = Font(size=10)
        ws['B3'].alignment = center_align

        # Insert Logo
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'image', 'college_logo.jpg')
        if os.path.exists(logo_path):
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 75
                img.height = 75
                ws.add_image(img, 'A1')
            except Exception as e:
                print(f"Logo error: {e}")

        # Title Row (Row 5-6)
        ws.merge_cells('A5:Q6')
        title_text = f"CONSOLIDATED INTERNAL MARKS - {subject.name.upper()} ({session.start_year.year}-{session.end_year.year})"
        ws['A5'] = title_text
        ws['A5'].font = title_font
        ws['A5'].alignment = center_align

        # Table Headers (Row 7 & 8)
        ws.merge_cells('A7:A8') # S.#
        ws.merge_cells('B7:B8') # ROLL NO
        ws.merge_cells('C7:C8') # STUDENT NAME
        ws.merge_cells('D7:H7') # MID-I
        ws.merge_cells('I7:M7') # MID-II
        ws.merge_cells('N7:P7') # CALCULATION
        ws.merge_cells('Q7:Q8') # Round Off

        headers = ['S.#', 'ROLL NO', 'STUDENT NAME', 'MID-TERM I (30M)', '', '', '', '', 'MID-TERM II (30M)', '', '', '', '', 'FINAL CALCULATION', '', '', 'FINAL']
        ws.append([''] * 17) # spacer for headers
        ws.append([''] * 17) # spacer for headers
        
        # Manually set header values to avoid appending mismatch
        ws['A7'], ws['B7'], ws['C7'], ws['Q7'] = 'S.#', 'ROLL NO', 'STUDENT NAME', 'FINAL'
        ws['D7'], ws['I7'], ws['N7'] = 'MID-TERM I (30M)', 'MID-TERM II (30M)', 'FINAL CALCULATION'
        
        sub_headers = ['','','','Obj','Dis','Asgn','Sum','Tot','Obj','Dis','Asgn','Sum','Tot','80%','20%','Total','']
        for c, text in enumerate(sub_headers, 1):
            if text: ws.cell(row=8, column=c, value=text)

        # Style Headers
        for r in [7, 8]:
            for c in range(1, 18):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(bold=True, color="FFFFFF" if r==7 else "000000")
                cell.fill = header_fill if r==7 else sub_header_fill
                cell.alignment = center_align
                cell.border = border

        # Data Rows (Starting Row 9)
        start_row = 9
        for i, student in enumerate(students, 1):
            # Fetch marks for Mid 1 and Mid 2
            # Use .order_by('id').last() to match web UI behavior of picking the most recent entry
            m1 = StudentResult.objects.filter(student=student, subject=subject, exam_name="Mid 1", semester=semester).order_by('id').last()
            m2 = StudentResult.objects.filter(student=student, subject=subject, exam_name="Mid 2", semester=semester).order_by('id').last()
            
            def get_vals(m):
                if not m: return {'obj':0, 'dis':0, 'asgn':0, 'sum':0, 'tot':0}
                dis_15 = round(m.descriptive / 2, 1) # Matches JS (m1.dis / 2).toFixed(1)
                m_sum = float(m.objective) + float(dis_15) + float(m.assignment)
                m_tot = min(js_round(m_sum), 30) # Matches JS Math.min(Math.round(m1_sum), 30)
                return {'obj':m.objective, 'dis':dis_15, 'asgn':m.assignment, 'sum':m_sum, 'tot':m_tot}
                
            v1 = get_vals(m1)
            v2 = get_vals(m2)
            
            best = max(v1['tot'], v2['tot'])
            other = min(v1['tot'], v2['tot'])
            p80 = round(best * 0.8, 2) # (best * 0.8).toFixed(2)
            p20 = round(other * 0.2, 2) # (other * 0.2).toFixed(2)
            raw_total = round(p80 + p20, 2) # (p80 + p20).toFixed(2)
            round_off = min(js_round(raw_total), 30) # Math.min(Math.round(raw_total), 30)
            
            # Name ordering: Last Name First Name
            full_name = f"{student.admin.last_name or ''} {student.admin.first_name or ''}".strip().upper()
            
            row_data = [
                i, student.roll_number, full_name,
                fmt_qty(v1['obj']), fmt_qty(v1['dis']), fmt_qty(v1['asgn']), fmt_qty(v1['sum']), fmt_qty(v1['tot']),
                fmt_qty(v2['obj']), fmt_qty(v2['dis']), fmt_qty(v2['asgn']), fmt_qty(v2['sum']), fmt_qty(v2['tot']),
                fmt_qty(p80), fmt_qty(p20), fmt_qty(raw_total), int(round_off)
            ]
            
            curr_row = start_row + i - 1
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=curr_row, column=c, value=val)
                cell.border = border
                cell.alignment = center_align if c != 3 else left_align
                if c in [8, 13, 17]: # Highlighting Total columns
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

        # Adjust columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        for c in range(4, 18):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 8

        # Footer space
        footer_row = start_row + len(students) + 2
        ws.merge_cells(f'C{footer_row}:E{footer_row}')
        ws[f'C{footer_row}'] = "SUBJECT TEACHER"
        ws[f'C{footer_row}'].font = Font(bold=True)
        
        ws.merge_cells(f'O{footer_row}:Q{footer_row}')
        ws[f'O{footer_row}'] = "HEAD OF DEPARTMENT"
        ws[f'O{footer_row}'].font = Font(bold=True)
        ws[f'O{footer_row}'].alignment = Alignment(horizontal='right')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"CONSOLIDATED_INTERNAL_{subject.name}_{session.start_year.year}.xlsx".replace(" ","_")
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return HttpResponse(f"Error generating Excel: {str(e)}")
